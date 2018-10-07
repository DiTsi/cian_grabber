from openpyxl import Workbook
from openpyxl import load_workbook
import copy
from time import sleep




first_ss = "downloads/offers_2018.10.01_00.48.34.xlsx"
second_ss = "downloads/offers_2018.10.07_21.35.07.xlsx"


def get_data_by_id(list, id, column):
    # found = False

    for i in range(len(list)):
        if list[i][0] == id:
            res = list[i][column]
            return res

    print("no id in list")
    return ""


def read_sheet(ws):
    l1 = list()
    l2 = list()
    l_res = list()
    empty = False

    for row in ws.rows:
        for cell in row:
            l2.append(cell.value)
        l1.append(copy.deepcopy(l2))
        l2.clear()
    for i in range(len(l1)):
        if len(l1[i]) == 0:
            empty = True
            break
    if empty:
        l_res = l1[0:i]
    else:
        l_res = l1
    return l_res


def get_new_ids(ws_list_first, ws_list_second):
    new_list = list()
    old_list = list()
    new_ids_list = list()

    for i in ws_list_second:
        new_list.append(i[0])

    for i in ws_list_first:
        old_list.append(i[0])

    for i in range(len(new_list)):
        if new_list[i] not in old_list:
            new_ids_list.append(new_list[i])

    sleep(1)

    return new_ids_list


def get_new_links(list, new_ids_list):
    counter = 1
    for i in range(len(new_ids_list)):
        for j in range(1, len(list)):
            if list[j][0] == new_ids_list[i]:
                print("{}. id: {}\t\tlink: {}".format(counter, new_ids_list[i], list[j][22]))
                counter += 1


first_wb = load_workbook(filename=first_ss, read_only=True)
second_wb = load_workbook(filename=second_ss, read_only=True)

first_ws = first_wb['ЦИАН - Продажа городской']
second_ws = second_wb['ЦИАН - Продажа городской']

l1 = read_sheet(first_ws)
l2 = read_sheet(second_ws)


e = get_new_ids(l1, l2)
get_new_links(l2, e)
# for i in e:
#     print(get_data_by_id(l2, i, 10) + "\n\n")

sleep(1)
#
# headers = ws.rows
# sleep(1)
# #
#

# print("l1 = " + len(first_ws.rows))

# cell = first_ws(row = 2, column = 2)
    # for cell in row:
    #     print(cell.value)


